import torch.cuda as cuda
import torch
import json
import os
from data_process import data_sampler
from transformers import BertTokenizer
from prototype_prompt import Protonets
from utils import *
from tqdm import tqdm
import pandas as pd
import torch.nn.functional as F
import openpyxl

workbook = openpyxl.Workbook()
worksheet = workbook.active


def eucli_tensor(x,y):
    return -1*torch.sqrt(torch.sum((x-y)*(x-y))).view(1)


if __name__ == "__main__":
    with open("config/config.json") as f:
        config = json.loads(f.read())
    device = torch.device('cuda', config["device_idx"])
    config['device'] = device
    config['n_gpu'] = torch.cuda.device_count()

    tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
    donum = 1
    max_len = 128
    hidden_layer_dim = 768

    for m in range(donum):

        sampler = data_sampler(config, tokenizer)
        sequence_results = []
        result_whole_test = []

        for ll in range(6):
            
            # Sample the number of categories from the private attribute "seed" in the sample class
            num_class = len(sampler.id2rel)
            set_seed(config, config['random_seed'] + 10 * ll)
            # By this step, a private attribute "seed" is written to the sample class
            sampler.set_seed(config['random_seed'] + 10 * ll)

            # Save all encountered relationships
            saveseen_relations = []
			# Prototype memory list
            test_data = []
            p2_map = {}
            log_data = f'./model/model.pkl'
            model = torch.load(log_data , map_location = config["device"])

            for steps, (training_data, valid_data, test_data, test_all_data, seen_relations, current_relations) in enumerate(sampler):
                worksheet.cell(row=1, column= steps+1,value=f'A{steps}')
                print(f"------- {str(steps)} stage started -------")
                prompt,proto = load_center(f'center/prompt_dict{steps}',f'center/proto_dict{steps}')

                savetest_all_data = test_all_data
                saveseen_relations = seen_relations
                currentnumber = len(current_relations)
                
                # Partition the data based on the category dictionary
                divide_train_set = {relation: [] for relation in current_relations}
                for data in training_data:
                    divide_train_set[data[0]].append(data)

                divide_test_set = {relation: [] for relation in seen_relations}
                for data in test_all_data:
                    if data[0] in seen_relations:
                        divide_test_set[data[0]].append(data)

                p2_map[steps] = current_relations
                
                test_accury = []
                proto_pool = []
                proto_pool.extend(proto[relation] for relation in seen_relations)
                query_set = []
                for label in seen_relations:
                    D_set = divide_test_set[label]
                    index_list = list(range(len(D_set)))
                    query = []
                    query.extend(D_set[i] for i in index_list)
                    query_set.append(query)
                
                for i in range(len(query_set)):
                    worksheet.cell(row=i+2, column=1, value=seen_relations[i])
                    data = []
                    index = []
                    t00 = []
                    for n in range(len(query_set[i])):
                        data.append(query_set[i][n][4])
                        index.append(query_set[i][n][5])

                    encode_input = tokenizer(data, return_tensors='pt',padding = True, max_length = 128, truncation = True)
                    encode_input['input_ids'] = torch.cat([torch.full((100,config["n_tokens"]), 0), encode_input['input_ids']], 1)
                    encode_input['attention_mask'] = torch.cat([torch.full((100,config["n_tokens"]), 1), encode_input['attention_mask']], 1)
                    encode_input['token_type_ids'] = torch.cat([torch.full((100,config["n_tokens"]), 0), encode_input['token_type_ids']], 1)
                    encode_input = encode_input.to(config["device"])
                    
                    featrue_q = [[0 for _ in range(len(prompt))] for _ in range(100)]
                    for k in range(len(prompt)):
                        layer_params_dict = model.state_dict()
                        layer_params_dict["embeddings.word_embeddings.learned_embedding"] = prompt[k]
                        model.load_state_dict(layer_params_dict)
                        featrue = model(**encode_input).last_hidden_state.detach()

                        for t in range(len(index)):
                            featrue_q[t][k] = featrue[t][index[t] + config['n_tokens']]
                    
                    for o in range(100):
                        for j in range(0,10*len(prompt)):
                            k = j//10
                            center_j = proto_pool[j]
                            if j == 0:
                                predict = eucli_tensor(featrue_q[o][k],center_j)
                            else:
                                predict = torch.cat((predict, eucli_tensor(featrue_q[o][k],center_j)), 0)
                        y_pre_j = int(torch.argmax(F.log_softmax(predict,dim=0)))
                        t00.append(1 if y_pre_j == i else 0)
                        test_accury.append(1 if y_pre_j == i else 0)
                    worksheet.cell(row=i+2, column=steps+2, value=sum(t00)/len(t00))
                test_acc = sum(test_accury)/len(test_accury)
                workbook.save(f'predict_{steps}.xlsx')
                print(f'The accuracy of step {str(steps)} is {test_acc}')
                
                
                
                
                
                    